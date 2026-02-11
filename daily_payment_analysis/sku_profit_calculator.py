#!/usr/bin/env python3

import pandas as pd
import sys
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------
# CONFIG
# ------------------------
SKU_PRICE_FILE = "data/sku_price.csv"
ORDER_FILE = "data/orders.xlsx"
ORDER_SHEET_NAME = "Order Payments"
OUTPUT_FILE = "data/sku_profit_report_v6.xlsx"

# ------------------------
# ADD PERCENTAGE COLUMNS
# ------------------------
def pct(numerator, denominator):
    if denominator == 0:
        return 0.0
    return round((numerator / denominator) * 100, 2)

# ------------------------
# LOAD FILES
# ------------------------
try:
    sku_prices = pd.read_csv(SKU_PRICE_FILE)

    # Load Excel with explicit sheet
    xls = pd.ExcelFile(ORDER_FILE)
    if ORDER_SHEET_NAME not in xls.sheet_names:
        print("❌ Sheet not found:", ORDER_SHEET_NAME)
        print("📄 Available sheets:")
        for s in xls.sheet_names:
            print(" -", s)
        sys.exit(1)

    orders = pd.read_excel(xls, sheet_name=ORDER_SHEET_NAME, header=1)

except Exception as e:
    print(f"❌ Error loading files: {e}")
    sys.exit(1)

# ------------------------
# NORMALIZE COLUMN NAMES
# ------------------------
def normalize_columns(df):
    df.columns = (
        df.columns
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    return df

sku_prices = normalize_columns(sku_prices)
orders = normalize_columns(orders)

# ------------------------
# REQUIRED COLUMNS
# ------------------------
required_price_cols = {"SKU", "buying_price"}
required_order_cols = {
    "Sub Order No",
    "Live Order Status",
    "Supplier SKU",
    "Final Settlement Amount",
    "Quantity"
}

missing_price = required_price_cols - set(sku_prices.columns)
missing_order = required_order_cols - set(orders.columns)

if missing_price:
    raise ValueError(f"❌ SKU price file missing columns: {missing_price}")

if missing_order:
    print("\n❌ Order file columns found:")
    for col in orders.columns:
        print(f" - {col}")
    raise ValueError(f"\n❌ Missing required order columns: {missing_order}")

# ------------------------
# PREPARE PRICE MAP
# ------------------------
sku_price_map = dict(
    zip(sku_prices["SKU"], sku_prices["buying_price"])
)

results = {}

# ------------------------
# PROCESS ORDERS
# ------------------------
for _, row in orders.iterrows():
    sku = row["Supplier SKU"]
    status = row["Live Order Status"]
    settlement = pd.to_numeric(
        str(row["Final Settlement Amount"]).replace(",", ""),
        errors="coerce"
    )

    if pd.isna(sku) or pd.isna(settlement):
        continue

    if sku not in sku_price_map:
        print(f"⚠️ SKU not found in price file: {sku}")
        continue

    qty = pd.to_numeric(row["Quantity"], errors="coerce")

    if pd.isna(qty) or qty <= 0:
        qty = 1

    if sku not in results:
        results[sku] = {
            "Total Orders": 0,
            "Delivered": 0,
            "Returned": 0,
            "RTO": 0,
            "Exchange": 0,
            "Revenue": 0.0,
            "Cost": 0.0,
            "Profit": 0.0
        }


    results[sku]["Total Orders"] += 1
    buying_price = sku_price_map[sku]

    cost = buying_price * qty

    # print("status: " + status +  " | sett: " + str(settlement))

    # --- FIX BLANK / NaN STATUS PROPERLY ---

    if pd.isna(status):
        status = ""
    else:
        status = str(status).strip()

    # If status is empty AND settlement exists → Delivered
    if status == "" and not pd.isna(settlement):
        status = "Delivered"


    # print("After: status: " + status +  " | sett: " + str(settlement))

    if status == "Delivered":
        results[sku]["Delivered"] += 1
        results[sku]["Revenue"] += settlement
        results[sku]["Cost"] += cost
        results[sku]["Profit"] += settlement - cost

    elif status == "Exchange":
        results[sku]["Exchange"] += 1
        results[sku]["Revenue"] += settlement
        results[sku]["Cost"] += cost
        results[sku]["Profit"] += settlement - cost

    elif status == "Return":
        results[sku]["Returned"] += 1
        results[sku]["Profit"] += settlement  # negative

    elif status == "RTO":
        results[sku]["RTO"] += 1
        continue

# ------------------------
# OUTPUT
# ------------------------
output_df = pd.DataFrame.from_dict(results, orient="index")
output_df.reset_index(inplace=True)
output_df.rename(columns={"index": "SKU"}, inplace=True)


output_df["Delivered %"] = output_df.apply(
    lambda r: pct(r["Delivered"] + r["Exchange"], r["Total Orders"]), axis=1
)

output_df["Returned %"] = output_df.apply(
    lambda r: pct(r["Returned"], r["Total Orders"]), axis=1
)

output_df["RTO %"] = output_df.apply(
    lambda r: pct(r["RTO"], r["Total Orders"]), axis=1
)


ordered_cols = [
    "SKU",
    "Total Orders",

    "Delivered",
    "Delivered %",

    "Exchange",

    "Returned",
    "Returned %",

    "RTO",
    "RTO %",

    "Revenue",
    "Cost",
    "Profit"
]

# Keep only columns that actually exist (extra safety)
ordered_cols = [c for c in ordered_cols if c in output_df.columns]

output_df = output_df[ordered_cols]

# output_df.to_csv(OUTPUT_FILE, index=False)

# print(f"\n✅ SKU-wise profit report generated: {OUTPUT_FILE}")


# ------------------------
# SAVE TO EXCEL FIRST
# ------------------------
output_df.to_excel(OUTPUT_FILE, index=False)

# ------------------------
# APPLY COLOR CODING
# ------------------------
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# Define fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
DARK_RED_FILL = PatternFill(start_color="EA9999", end_color="EA9999", fill_type="solid")

# Find Returned % column index
header = [cell.value for cell in ws[1]]
returned_pct_col = header.index("Returned %") + 1

# Apply row-wise coloring
for row in range(2, ws.max_row + 1):
    returned_pct = ws.cell(row=row, column=returned_pct_col).value

    if returned_pct is None:
        continue

    if returned_pct <= 10:
        fill = GREEN_FILL
    elif returned_pct > 30:
        fill = DARK_RED_FILL
    elif returned_pct > 20:
        fill = LIGHT_RED_FILL
    else:
        continue  # neutral

    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = fill

# Save final workbook
wb.save(OUTPUT_FILE)

print(f"✅ Excel report generated with color coding: {OUTPUT_FILE}")

