#!/usr/bin/env python3

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================
# CONFIG
# ============================================================
SKU_PRICE_FILE = "data/sku_price.csv"
ORDER_FILE = "data/orders.xlsx"
ORDER_SHEET_NAME = "Order Payments"
OUTPUT_FILE = "data/sku_profit_report_v9.xlsx"

# ============================================================
# HELPERS
# ============================================================
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = (
        df.columns.astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    return df

def parse_money(value):
    return pd.to_numeric(str(value).replace(",", ""), errors="coerce")

def pct(numerator, denominator) -> float:
    if denominator == 0:
        return 0.0
    return round((numerator / denominator) * 100, 2)


def normalize_status(raw_status, settlement):
    """
    Rules:
    - NaN / empty status + settlement present → Delivered
    """
    if pd.isna(raw_status):
        status = ""
    else:
        status = str(raw_status).strip()

    if status == "" and not pd.isna(settlement):
        return "Delivered"

    return status


def parse_settlement(value):
    return pd.to_numeric(str(value).replace(",", ""), errors="coerce")


def parse_quantity(value):
    qty = pd.to_numeric(value, errors="coerce")
    if pd.isna(qty) or qty <= 0:
        return 1
    return qty


# ============================================================
# LOAD DATA
# ============================================================
try:
    sku_prices = normalize_columns(pd.read_csv(SKU_PRICE_FILE))

    xls = pd.ExcelFile(ORDER_FILE)
    if ORDER_SHEET_NAME not in xls.sheet_names:
        print("❌ Sheet not found:", ORDER_SHEET_NAME)
        print("📄 Available sheets:", xls.sheet_names)
        sys.exit(1)

    orders = normalize_columns(
        pd.read_excel(xls, sheet_name=ORDER_SHEET_NAME, header=1)
    )

except Exception as e:
    print(f"❌ Error loading files: {e}")
    sys.exit(1)

# ============================================================
# VALIDATION
# ============================================================
required_price_cols = {"SKU", "buying_price"}
required_order_cols = {
    "Sub Order No",
    "Live Order Status",
    "Supplier SKU",
    "Final Settlement Amount",
    "Quantity",
}

if not required_price_cols.issubset(sku_prices.columns):
    raise ValueError(f"❌ SKU price file missing columns: {required_price_cols}")

if not required_order_cols.issubset(orders.columns):
    print("\n❌ Order file columns found:")
    for col in orders.columns:
        print(" -", col)
    raise ValueError("❌ Order file missing required columns")

# ============================================================
# PREP DATA
# ============================================================
sku_price_map = dict(zip(sku_prices["SKU"], sku_prices["buying_price"]))

results = {}

# ============================================================
# PROCESS ORDERS
# ============================================================
for _, row in orders.iterrows():
    sku = row["Supplier SKU"]
    settlement = parse_settlement(row["Final Settlement Amount"])

    if pd.isna(sku) or pd.isna(settlement):
        continue

    if sku not in sku_price_map:
        print(f"⚠️ SKU not found in price file: {sku}")
        continue

    qty = parse_quantity(row["Quantity"])
    status = normalize_status(row["Live Order Status"], settlement)

    if sku not in results:
        results[sku] = {
            "Total Orders": 0,
            "Delivered": 0,
            "Exchange": 0,
            "Returned": 0,
            "RTO": 0,
            "Revenue": 0.0,
            "Cost": 0.0,
            "Profit": 0.0,
        }

    results[sku]["Total Orders"] += 1

    buying_price = sku_price_map[sku]
    cost = buying_price * qty

    if status == "Delivered":
        results[sku]["Delivered"] += 1
        results[sku]["Revenue"] += settlement
        results[sku]["Cost"] += cost
        results[sku]["Profit"] += settlement - cost

    elif status == "Exchange":
        results[sku]["Exchange"] += 1
        results[sku]["Revenue"] += settlement
        results[sku]["Cost"] += cost
        results[sku]["Profit"] += settlement - cost

    elif status == "Return":
        results[sku]["Returned"] += 1
        results[sku]["Profit"] += settlement

    elif status == "RTO":
        results[sku]["RTO"] += 1

# ============================================================
# BUILD OUTPUT
# ============================================================
output_df = (
    pd.DataFrame.from_dict(results, orient="index")
    .reset_index()
    .rename(columns={"index": "SKU"})
)

output_df["Delivered %"] = output_df.apply(
    lambda r: pct(r["Delivered"] + r["Exchange"], r["Total Orders"]), axis=1
)
output_df["Returned %"] = output_df.apply(
    lambda r: pct(r["Returned"], r["Total Orders"]), axis=1
)
output_df["RTO %"] = output_df.apply(
    lambda r: pct(r["RTO"], r["Total Orders"]), axis=1
)

ordered_cols = [
    "SKU",
    "Total Orders",
    "Delivered",
    "Delivered %",
    "Exchange",
    "Returned",
    "Returned %",
    "RTO",
    "RTO %",
    "Revenue",
    "Cost",
    "Profit",
]

output_df = output_df[[c for c in ordered_cols if c in output_df.columns]]


# ============================================================
# ADS COST CALCULATION
# ============================================================
try:
    ads_df = normalize_columns(
        pd.read_excel(
            ORDER_FILE,
            sheet_name="Ads Cost",
            header=1,
            skiprows=[2]
        )
    )

    if "Total Ads Cost" not in ads_df.columns:
        raise ValueError("❌ 'Total Ads Cost' column not found in Ads Cost sheet")

    ads_df["Total Ads Cost"] = ads_df["Total Ads Cost"].apply(parse_money)

    total_ads_cost = ads_df["Total Ads Cost"].sum(skipna=True)

except Exception as e:
    print(f"⚠️ Ads cost not applied: {e}")
    total_ads_cost = 0.0


# ============================================================
# OVERALL PROFIT
# ============================================================
overall_product_profit = output_df["Profit"].sum()
net_profit_after_ads = overall_product_profit + total_ads_cost


# ============================================================
# SAVE EXCEL
# ============================================================
# output_df.to_excel(OUTPUT_FILE, index=False)



# ============================================================
# ADD SUMMARY SHEET
# ============================================================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    output_df.to_excel(writer, sheet_name="SKU Report", index=False)

    summary_df = pd.DataFrame({
        "Metric": [
            "Overall Product Profit",
            "Total Ads Cost",
            "Net Profit After Ads"
        ],
        "Amount": [
            round(overall_product_profit, 2),
            round(total_ads_cost, 2),
            round(net_profit_after_ads, 2)
        ]
    })

    summary_df.to_excel(writer, sheet_name="Summary", index=False)


# ============================================================
# COLOR CODING (Returned %)
# ============================================================
wb = load_workbook(OUTPUT_FILE)
ws = wb["SKU Report"]

GREEN = PatternFill("solid", fgColor="C6EFCE")
LIGHT_RED = PatternFill("solid", fgColor="F4CCCC")
DARK_RED = PatternFill("solid", fgColor="EA9999")

headers = [c.value for c in ws[1]]
returned_pct_col = headers.index("Returned %") + 1

for r in range(2, ws.max_row + 1):
    val = ws.cell(r, returned_pct_col).value
    if val is None:
        continue

    if val <= 10:
        fill = GREEN
    elif val > 30:
        fill = DARK_RED
    elif val > 20:
        fill = LIGHT_RED
    else:
        continue

    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).fill = fill

wb.save(OUTPUT_FILE)

print(f"✅ Excel report generated with color coding: {OUTPUT_FILE}")
