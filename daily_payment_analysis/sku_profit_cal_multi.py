#!/usr/bin/env python3

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================
# CONFIG
# ============================================================
BASE_DIR = Path("data")
SKU_PRICE_FILE = BASE_DIR / "sku_price.csv"
ORDERS_DIR = BASE_DIR / "orders"
OUTPUT_FILE = BASE_DIR / "sku_profit_report_final.xlsx"

ORDER_SHEET_NAME = "Order Payments"
ADS_SHEET_NAME = "Ads Cost"

BASE_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# HELPERS (BUSINESS LOGIC UNCHANGED)
# ============================================================
def normalize_columns(df):
    df.columns = (
        df.columns.astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    return df


def parse_money(val):
    return pd.to_numeric(str(val).replace(",", ""), errors="coerce")


def parse_qty(val):
    q = pd.to_numeric(val, errors="coerce")
    return 1 if pd.isna(q) or q <= 0 else q


def normalize_status(raw_status, settlement):
    if pd.isna(raw_status):
        status = ""
    else:
        status = str(raw_status).strip()

    if status == "" and not pd.isna(settlement):
        return "Delivered"

    return status


# ============================================================
# LOAD SKU PRICES
# ============================================================
sku_prices = normalize_columns(pd.read_csv(SKU_PRICE_FILE))
sku_price_map = dict(zip(sku_prices["SKU"], sku_prices["buying_price"]))

# ============================================================
# COLLECT RESULTS (THIS IS THE FIX)
# ============================================================
all_sku_dfs = []
summaries = []

# ============================================================
# PROCESS EACH ORDER FILE
# ============================================================
for order_file in sorted(ORDERS_DIR.glob("*.xlsx")):
    print(f"▶ Processing: {order_file.name}")

    orders = normalize_columns(
        pd.read_excel(order_file, sheet_name=ORDER_SHEET_NAME, header=1)
    )

    file_rows = []


    # ---------------- Order Processing ----------------
    for _, row in orders.iterrows():
        sku = row.get("Supplier SKU")
        settlement = parse_money(row.get("Final Settlement Amount"))

        print("sku" + sku)
        print("settlemtn" + settlement)
        
        if pd.isna(sku) or pd.isna(settlement):
            continue

        if sku not in sku_price_map:
            continue

        qty = parse_qty(row.get("Quantity"))
        status = normalize_status(row.get("Live Order Status"), settlement)
        cost = sku_price_map[sku] * qty

        print(qty)

        record = {
            "SKU": sku,
            "Total Orders": 1,
            "Delivered": 0,
            "Exchange": 0,
            "Returned": 0,
            "RTO": 0,
            "Revenue": 0.0,
            "Cost": 0.0,
            "Profit": 0.0,
        }

        print("1")

        if status == "Delivered":
            record["Delivered"] = 1
            record["Revenue"] = settlement
            record["Cost"] = cost
            record["Profit"] = settlement - cost

        elif status == "Exchange":
            record["Exchange"] = 1
            record["Revenue"] = settlement
            record["Cost"] = cost
            record["Profit"] = settlement - cost

        elif status == "Return":
            record["Returned"] = 1
            record["Profit"] = settlement  # already negative

        elif status == "RTO":
            record["RTO"] = 1

        file_rows.append(record)

    if not file_rows:
        print(f"⚠️ No valid rows in {order_file.name}")
        continue

    # Freeze THIS FILE
    file_df = pd.DataFrame(file_rows)

    # SKU-wise aggregation for this file
    file_sku_df = (
        file_df.groupby("SKU", as_index=False)
        .sum()
    )

    all_sku_dfs.append(file_sku_df)

    # ---------------- Ads Cost ----------------
    ads_df = normalize_columns(
        pd.read_excel(
            order_file,
            sheet_name=ADS_SHEET_NAME,
            header=1,
            skiprows=[2]
        )
    )

    ads_df["Total Ads Cost"] = ads_df["Total Ads Cost"].apply(parse_money)
    total_ads_cost = ads_df["Total Ads Cost"].sum(skipna=True)

    product_profit = file_sku_df["Profit"].sum()
    net_profit = product_profit + total_ads_cost

    summaries.append({
        "file": order_file.stem,
        "product_profit": round(product_profit, 2),
        "ads_cost": round(total_ads_cost, 2),
        "net_profit": round(net_profit, 2),
    })

# ============================================================
# FINAL SKU CONSOLIDATION (NO OVERRIDE POSSIBLE)
# ============================================================
if not all_sku_dfs:
    raise SystemExit("❌ No SKU data found across files.")

sku_df = (
    pd.concat(all_sku_dfs, ignore_index=True)
    .groupby("SKU", as_index=False)
    .sum()
)

# ============================================================
# PERCENTAGES (STABLE)
# ============================================================
sku_df["Delivered %"] = (
    (sku_df["Delivered"] + sku_df["Exchange"]) / sku_df["Total Orders"]
).fillna(0).mul(100).round(2)

sku_df["Returned %"] = (
    sku_df["Returned"] / sku_df["Total Orders"]
).fillna(0).mul(100).round(2)

sku_df["RTO %"] = (
    sku_df["RTO"] / sku_df["Total Orders"]
).fillna(0).mul(100).round(2)

sku_df = sku_df[
    [
        "SKU", "Total Orders",
        "Delivered", "Delivered %",
        "Exchange",
        "Returned", "Returned %",
        "RTO", "RTO %",
        "Revenue", "Cost", "Profit",
    ]
]

# ============================================================
# WRITE EXCEL
# ============================================================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    sku_df.to_excel(writer, sheet_name="SKU Report", index=False)

    for s in summaries:
        pd.DataFrame({
            "Metric": [
                "Overall Product Profit",
                "Total Ads Cost",
                "Net Profit After Ads",
            ],
            "Amount": [
                s["product_profit"],
                s["ads_cost"],
                s["net_profit"],
            ]
        }).to_excel(writer, sheet_name=f"Summary_{s['file']}", index=False)

# ============================================================
# COLOR CODING (Returned %)
# ============================================================
wb = load_workbook(OUTPUT_FILE)
ws = wb["SKU Report"]

GREEN = PatternFill("solid", fgColor="C6EFCE")
LIGHT_RED = PatternFill("solid", fgColor="F4CCCC")
DARK_RED = PatternFill("solid", fgColor="EA9999")

headers = [c.value for c in ws[1]]
ret_col = headers.index("Returned %") + 1

for r in range(2, ws.max_row + 1):
    val = ws.cell(r, ret_col).value
    if val is None:
        continue

    if val <= 8:
        fill = GREEN
    elif val > 25:
        fill = DARK_RED
    else:
        continue

    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).fill = fill

wb.save(OUTPUT_FILE)

print(f"\n✅ FINAL REPORT GENERATED: {OUTPUT_FILE}")