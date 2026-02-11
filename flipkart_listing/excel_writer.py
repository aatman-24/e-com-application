# excel_writer.py

from openpyxl import load_workbook

def write_rows_to_excel(excel_path, sheet_name, rows):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # Map headers
    column_map = {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }

    sku_col = column_map["Seller SKU ID"]

    # Find first empty row
    row_ptr = None
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, sku_col).value:
            row_ptr = r
            break
    if row_ptr is None:
        row_ptr = ws.max_row + 1

    # Write rows
    for row in rows:
        for key, val in row.items():
            if key in column_map:
                ws.cell(row_ptr, column_map[key]).value = val
        row_ptr += 1

    wb.save(excel_path)
